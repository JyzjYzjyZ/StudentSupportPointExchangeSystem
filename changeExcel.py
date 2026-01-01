import openpyxl
from openpyxl import Workbook
import os
import sys

# 读取Excel文件 - 接受命令行参数作为文件路径
if len(sys.argv) < 2:
    print("请提供Excel文件路径作为参数")
    sys.exit(1)

excel_path = sys.argv[1]
workbook = openpyxl.load_workbook(excel_path)

# 创建新的工作簿来存储合并后的数据
merged_workbook = Workbook()
merged_sheet = merged_workbook.active
merged_sheet.title = '合并数据'

# 获取所有sheet名称
sheet_names = workbook.sheetnames

# 遍历每个sheet
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    
    # 获取当前sheet的所有行
    rows = list(sheet.iter_rows(values_only=True))
    
    # 如果是第一个sheet，复制所有行（包括表头）
    if sheet_name == sheet_names[0]:
        for row in rows:
            merged_sheet.append(row)
    # 否则，跳过表头，只复制数据行
    else:
        for row in rows[1:]:
            merged_sheet.append(row)

# 删除第一行（标题行）
merged_sheet.delete_rows(1)

# 确保第一行是正确的表头
correct_headers = ["姓名", "学号", "学院", "爱心币数量", "剩余爱心币"]
# 获取当前第一行
current_header = merged_sheet[1]
# 检查当前表头是否正确
header_mismatch = False
for i, (current_cell, correct_header) in enumerate(zip(current_header, correct_headers)):
    if current_cell.value != correct_header:
        header_mismatch = True
        break

# 如果表头不正确，替换为正确的表头
if header_mismatch:
    for i, header in enumerate(correct_headers, 1):
        merged_sheet.cell(row=1, column=i).value = header

# 处理D列和E列的数据
# 从第3行开始处理，跳过第1、2行
for row in range(3, merged_sheet.max_row + 1):
    # 获取D列和E列的值
    d_value = merged_sheet[f'D{row}'].value
    e_value = merged_sheet[f'E{row}'].value
    
    # 确定最终值：如果E列有数值（不是None），则使用E列的值，否则使用D列的值
    final_value = e_value if e_value is not None else d_value
    
    # 将最终值写入D列（爱心币数量）
    merged_sheet[f'D{row}'] = final_value
    # 保留E列（剩余爱心币）的原始值，不做修改

# 删除空白行（从下往上检查，避免删除行后索引混乱）
# 从最后一行开始往上检查
for row in range(merged_sheet.max_row, 2, -1):  # 从最后一行到第3行（保留第1、2行）
    # 检查当前行是否所有单元格都是空的
    is_empty = True
    for col in range(1, merged_sheet.max_column + 1):
        cell_value = merged_sheet.cell(row=row, column=col).value
        if cell_value is not None and cell_value != '':
            is_empty = False
            break
    # 如果是空白行，则删除
    if is_empty:
        merged_sheet.delete_rows(row)

# 删除包含"高等职业技术学院"的行（从下往上检查，避免删除行后索引混乱）
delete_count = 0
for row in range(merged_sheet.max_row, 2, -1):  # 从最后一行到第3行（保留第1、2行）
    # 检查当前行的所有单元格
    for col in range(1, merged_sheet.max_column + 1):
        cell_value = merged_sheet.cell(row=row, column=col).value
        if cell_value is not None and "高等职业技术学院" in str(cell_value):
            merged_sheet.delete_rows(row)
            delete_count += 1
            break

# 保存合并后的文件
# 生成输出文件路径：在输入文件目录下生成处理后的文件
input_dir = os.path.dirname(excel_path)
input_filename = os.path.basename(excel_path)
name, ext = os.path.splitext(input_filename)
merged_file_path = os.path.join(input_dir, f'{name}_处理后{ext}')
merged_workbook.save(merged_file_path)

print(f'Excel文件合并完成！合并后的文件保存在：{merged_file_path}')
print(f'原始文件包含 {len(sheet_names)} 个sheet')
print(f'合并后的数据总行数：{merged_sheet.max_row}')
print(f'已完成D列（爱心币数量）的数据处理，E列（剩余爱心币）保持不变')
print(f'已删除包含"高等职业技术学院"的行：{delete_count} 行')