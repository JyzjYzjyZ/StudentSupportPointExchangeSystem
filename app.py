from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
from werkzeug.utils import secure_filename
import uuid  # 用于生成唯一文件名
from openpyxl import load_workbook  # 新增：处理Excel文件
import sqlite3  # 新增：直接使用sqlite3
import tempfile  # 新增：处理临时文件
import io  # 新增：字节流处理

app = Flask(__name__)
app.secret_key = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///shop.db'
app.config['UPLOAD_FOLDER'] = 'static/images'  # 图片上传目录
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}  # 允许的图片格式

# 确保上传目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)


class PurchaseRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    purchase_time = db.Column(db.DateTime, default=datetime.utcnow)

    # 关系
    product = db.relationship('Product', backref='purchases')
    user = db.relationship('User', backref='purchases')


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    picture = db.Column(db.String(200), nullable=False)
    price = db.Column(db.Float, nullable=False)
    stock = db.Column(db.Integer, nullable=False)
    limit = db.Column(db.Integer, nullable=False)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(80), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    name = db.Column(db.String(80))  # 姓名
    gender = db.Column(db.String(10))  # 性别
    college = db.Column(db.String(100))  # 新增：学院
    points = db.Column(db.Integer, default=0)  # 积分（对应爱心币数量）
    remaining_points = db.Column(db.Integer, default=0)  # 新增：剩余爱心币


# 辅助函数：检查文件是否为允许的图片格式
def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# 修复后的数据库表结构更新函数
def update_database_schema():
    # 直接使用sqlite3连接数据库
    conn = sqlite3.connect('shop.db')
    cursor = conn.cursor()

    # 检查并添加college字段
    try:
        cursor.execute('ALTER TABLE user ADD COLUMN college VARCHAR(100)')
        print("成功添加college字段")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e).lower():
            print("college字段已存在")
        else:
            print(f"添加college字段出错: {e}")

    # 检查并添加remaining_points字段
    try:
        cursor.execute('ALTER TABLE user ADD COLUMN remaining_points INTEGER DEFAULT 0')
        print("成功添加remaining_points字段")
    except sqlite3.OperationalError as e:
        if "duplicate column name" in str(e).lower():
            print("remaining_points字段已存在")
        else:
            print(f"添加remaining_points字段出错: {e}")

    conn.commit()
    conn.close()


@app.route('/')
def home():
    if 'username' not in session:
        return redirect(url_for('login'))

    if session.get('is_admin'):
        return redirect(url_for('admin'))
    else:
        products = Product.query.all()
        user = User.query.filter_by(username=session['username']).first()
        purchase_records = PurchaseRecord.query \
            .filter_by(user_id=user.id) \
            .join(Product) \
            .order_by(PurchaseRecord.purchase_time.desc()) \
            .all()
        return render_template('shop.html',
                               products=products,
                               current_points=user.points,
                               purchase_records=purchase_records)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.query.filter_by(username=username, password=password).first()
        if user:
            session['username'] = username
            session['is_admin'] = user.is_admin
            return redirect(url_for('home'))
        else:
            return render_template('login.html', error='Invalid credentials')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin')
def admin():
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    products = Product.query.all()
    users = User.query.all()
    return render_template('admin.html', products=products, users=users)


@app.route('/add_user', methods=['POST'])
def add_user():
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))

    username = request.form['username']
    password = request.form['password']
    is_admin = 'is_admin' in request.form
    name = request.form.get('name', '')
    gender = request.form.get('gender', 'male')
    points = int(request.form.get('points', 0))

    new_user = User(
        username=username,
        password=password,
        is_admin=is_admin,
        name=name,
        gender=gender,
        points=points
    )
    db.session.add(new_user)
    db.session.commit()

    return redirect(url_for('admin'))


@app.route('/add_product', methods=['POST'])
def add_product():
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    if 'picture' not in request.files:
        return "未选择图片", 400
    file = request.files['picture']
    if file.filename == '':
        return "未选择图片", 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        name = request.form['name']
        price = float(request.form['price'])
        stock = int(request.form['stock'])
        limit = int(request.form['limit'])

        new_product = Product(
            name=name,
            picture=unique_filename,
            price=price,
            stock=stock,
            limit=limit
        )
        db.session.add(new_product)
        db.session.commit()

        return redirect(url_for('admin'))

    return "不支持的文件格式", 400


@app.route('/purchase/<int:product_id>', methods=['POST'])
def purchase(product_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    product = Product.query.get_or_404(product_id)
    user = User.query.filter_by(username=session['username']).first()
    quantity = int(request.form['quantity'])
    total_cost = product.price * quantity

    # 检查本月购买记录
    now = datetime.utcnow()
    first_day_of_month = datetime(now.year, now.month, 1)
    purchased_this_month = PurchaseRecord.query.filter(
        PurchaseRecord.user_id == user.id,
        PurchaseRecord.product_id == product.id,
        PurchaseRecord.purchase_time >= first_day_of_month
    ).with_entities(db.func.sum(PurchaseRecord.quantity)).scalar() or 0

    if purchased_this_month + quantity > product.limit:
        return f"本月已购买{purchased_this_month}件，超过限购数量", 400

    if quantity > product.limit or quantity > product.stock:
        return "超过购买限制或库存不足", 400
    if user.points < total_cost:
        return "积分不足", 400

    # 扣减库存和用户积分
    product.stock -= quantity
    user.points -= total_cost
    user.remaining_points = user.points  # 更新剩余爱心币
    db.session.commit()

    # 记录购买
    record = PurchaseRecord(
        user_id=user.id,
        product_id=product.id,
        quantity=quantity
    )
    db.session.add(record)
    db.session.commit()

    return render_template('purchase_success.html',
                           product_name=product.name,
                           quantity=quantity,
                           total_cost=total_cost,
                           current_points=user.points)


@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    student = User.query.get(student_id)
    if student and not student.is_admin:  # 确保不是管理员
        db.session.delete(student)
        db.session.commit()
    return redirect(url_for('admin'))


@app.route('/increase_stock/<int:product_id>', methods=['POST'])
def increase_stock(product_id):
    product = Product.query.get_or_404(product_id)
    product.stock += 1
    db.session.commit()
    return redirect(url_for('admin'))


@app.route('/decrease_stock/<int:product_id>', methods=['POST'])
def decrease_stock(product_id):
    product = Product.query.get_or_404(product_id)
    if product.stock > 0:
        product.stock -= 1
        db.session.commit()
    return redirect(url_for('admin'))


@app.route('/get_product/<int:product_id>')
def get_product(product_id):
    product = Product.query.get_or_404(product_id)
    return {
        'name': product.name,
        'picture': product.picture,
        'price': float(product.price),
        'stock': product.stock,
        'limit': product.limit
    }


@app.route('/update_product/<int:product_id>', methods=['POST'])
def update_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.name = request.form['name']
    product.price = float(request.form['price'])
    product.stock = int(request.form['stock'])
    product.limit = int(request.form['limit'])
    if 'picture' in request.files and request.files['picture'].filename != '':
        file = request.files['picture']
        if file and allowed_file(file.filename):
            old_picture_path = os.path.join(app.config['UPLOAD_FOLDER'], product.picture)
            if os.path.exists(old_picture_path):
                os.remove(old_picture_path)
            filename = secure_filename(file.filename)
            unique_filename = f"{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            product.picture = unique_filename

    db.session.commit()
    return '', 204


@app.route('/delete_admin/<int:admin_id>', methods=['POST'])
def delete_admin(admin_id):
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))

    admin = User.query.get_or_404(admin_id)
    # 防止删除最后一个管理员
    admin_count = User.query.filter_by(is_admin=True).count()
    if admin.is_admin and admin_count > 1:
        db.session.delete(admin)
        db.session.commit()
        return redirect(url_for('admin'))
    return "不能删除最后一个管理员", 403


@app.route('/get_user/<int:user_id>')
def get_user(user_id):
    user = User.query.get_or_404(user_id)
    return {
        'username': user.username,
        'name': user.name,
        'gender': user.gender,
        'college': user.college,
        'points': user.points,
        'remaining_points': user.remaining_points,
        'password': user.password
    }


@app.route('/update_user/<int:user_id>', methods=['POST'])
def update_user(user_id):
    user = User.query.get_or_404(user_id)
    user.username = request.form['username']
    user.name = request.form['name']
    user.gender = request.form['gender']
    user.points = int(request.form['points'])
    user.remaining_points = user.points  # 同步更新剩余爱心币
    # 更新学院字段
    if 'college' in request.form:
        user.college = request.form['college']
    # 更新密码字段
    if 'password' in request.form and request.form['password']:
        user.password = request.form['password']
    db.session.commit()
    return '', 204


@app.route('/return_purchase/<int:record_id>', methods=['POST'])
def return_purchase(record_id):
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))

    record = PurchaseRecord.query.get_or_404(record_id)
    user = User.query.get_or_404(record.user_id)
    product = Product.query.get_or_404(record.product_id)

    # 返还积分和库存
    refund_amount = record.quantity * product.price
    user.points += refund_amount
    user.remaining_points = user.points  # 同步更新剩余爱心币
    product.stock += record.quantity

    # 删除购买记录
    db.session.delete(record)
    db.session.commit()

    return '', 204


# 修复后的Excel导入路由
@app.route('/import_excel', methods=['POST'])
def import_excel():
    if 'username' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))

    if 'excel_file' not in request.files:
        return "未选择文件", 400

    file = request.files['excel_file']
    if file.filename == '':
        return "未选择文件", 400

    # 检查文件格式
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return "请上传Excel文件", 400

    try:
        # 修复：将上传的文件转换为字节流，解决seekable问题
        # 方法1：读取文件内容到字节流
        file_content = file.read()
        file_stream = io.BytesIO(file_content)

        # 加载Excel文件（使用字节流）
        workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook.active

        # 获取表头并验证
        headers = [cell.value for cell in sheet[1]]
        required_headers = ['姓名', '学号', '学院', '爱心币数量', '剩余爱心币']
        for header in required_headers:
            if header not in headers:
                return f"Excel文件缺少必要列: {header}", 400

        # 记录导入成功和失败的数量
        success_count = 0
        fail_count = 0

        # 从第二行开始读取数据（跳过表头）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 创建数据字典，将单元格值与表头对应
            row_data = dict(zip(headers, row))

            # 提取所需数据
            name = row_data['姓名'] or ''
            student_id = row_data['学号'] or ''
            college = row_data['学院'] or ''
            points = int(row_data['爱心币数量'] or 0)
            remaining_points = int(row_data['剩余爱心币'] or 0)

            # 检查学号是否为空
            if not student_id:
                fail_count += 1
                continue

            # 检查用户是否已存在
            existing_user = User.query.filter_by(username=str(student_id)).first()
            if existing_user:
                # 更新现有用户信息
                existing_user.name = name
                existing_user.college = college
                existing_user.points = points
                existing_user.remaining_points = remaining_points
                success_count += 1
            else:
                # 创建新用户（密码默认为学号）
                new_user = User(
                    username=str(student_id),
                    password=str(student_id),
                    is_admin=False,
                    name=name,
                    college=college,
                    points=points,
                    remaining_points=remaining_points,
                    gender='male'  # 默认性别
                )
                db.session.add(new_user)
                success_count += 1

        db.session.commit()
        return f"导入成功！成功 {success_count} 条，失败 {fail_count} 条", 200

    except Exception as e:
        db.session.rollback()
        return f"导入失败: {str(e)}", 500


@app.route('/deepseek')
def deepseek():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('deepseek.html')


@app.route('/help')
def help_page():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('help.html')


@app.route('/process_excel', methods=['POST'])
def process_excel():
    if 'username' not in session or not session.get('is_admin'):
        return jsonify({'success': False, 'error': '未授权访问'}), 403

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    # 检查文件格式
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': '请上传Excel文件'}), 400

    try:
        # 创建临时目录保存上传的文件
        temp_dir = os.path.join(app.root_path, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        # 保存文件到临时目录
        temp_file_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(temp_file_path)
        
        # 直接在Flask应用中处理Excel文件
        import openpyxl
        from openpyxl import Workbook
        
        # 读取Excel文件
        workbook = openpyxl.load_workbook(temp_file_path)
        
        # 创建新的工作簿来存储合并后的数据
        merged_workbook = Workbook()
        merged_sheet = merged_workbook.active
        merged_sheet.title = '合并数据'
        
        # 获取所有sheet名称
        sheet_names = workbook.sheetnames
        original_sheets = len(sheet_names)
        
        # 遍历每个sheet
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            
            # 获取当前sheet的所有行
            rows = list(sheet.iter_rows(values_only=True))
            
            # 如果是第一个sheet，复制所有行（包括表头）
            if sheet_name == sheet_names[0]:
                for row in rows:
                    merged_sheet.append(row)
            # 否则，跳过表头，只复制数据行
            else:
                for row in rows[1:]:
                    merged_sheet.append(row)
        
        # 删除第一行（标题行）
        merged_sheet.delete_rows(1)
        
        # 确保第一行是正确的表头
        correct_headers = ["姓名", "学号", "学院", "爱心币数量", "剩余爱心币"]
        # 获取当前第一行
        current_header = merged_sheet[1]
        # 检查当前表头是否正确
        header_mismatch = False
        for i, (current_cell, correct_header) in enumerate(zip(current_header, correct_headers)):
            if current_cell.value != correct_header:
                header_mismatch = True
                break
        
        # 如果表头不正确，替换为正确的表头
        if header_mismatch:
            for i, header in enumerate(correct_headers, 1):
                merged_sheet.cell(row=1, column=i).value = header
        
        # 处理D列和E列的数据
        # 从第3行开始处理，跳过第1、2行
        for row in range(3, merged_sheet.max_row + 1):
            # 获取D列和E列的值
            d_value = merged_sheet[f'D{row}'].value
            e_value = merged_sheet[f'E{row}'].value
            
            # 确定最终值：如果E列有数值（不是None），则使用E列的值，否则使用D列的值
            final_value = e_value if e_value is not None else d_value
            
            # 将最终值写入D列（爱心币数量）
            merged_sheet[f'D{row}'] = final_value
            # 保留E列（剩余爱心币）的原始值，不做修改
        
        # 删除空白行（从下往上检查，避免删除行后索引混乱）
        # 从最后一行开始往上检查
        for row in range(merged_sheet.max_row, 2, -1):  # 从最后一行到第3行（保留第1、2行）
            # 检查当前行是否所有单元格都是空的
            is_empty = True
            for col in range(1, merged_sheet.max_column + 1):
                cell_value = merged_sheet.cell(row=row, column=col).value
                if cell_value is not None and cell_value != '':
                    is_empty = False
                    break
            # 如果是空白行，则删除
            if is_empty:
                merged_sheet.delete_rows(row)
        
        # 删除包含"高等职业技术学院"的行（从下往上检查，避免删除行后索引混乱）
        delete_count = 0
        for row in range(merged_sheet.max_row, 2, -1):  # 从最后一行到第3行（保留第1、2行）
            # 检查当前行的所有单元格
            for col in range(1, merged_sheet.max_column + 1):
                cell_value = merged_sheet.cell(row=row, column=col).value
                if cell_value is not None and "高等职业技术学院" in str(cell_value):
                    merged_sheet.delete_rows(row)
                    delete_count += 1
                    break
        
        # 创建Excel输出目录
        excel_output_dir = os.path.join(app.root_path, 'static', 'excel')
        os.makedirs(excel_output_dir, exist_ok=True)
        
        # 保存合并后的文件
        # 生成输出文件路径：在静态目录下生成处理后的文件
        input_filename = os.path.basename(temp_file_path)
        name, ext = os.path.splitext(input_filename)
        merged_filename = f'{name}_处理后{ext}'
        merged_file_path = os.path.join(excel_output_dir, merged_filename)
        merged_workbook.save(merged_file_path)
        
        # 生成前端可访问的URL路径
        merged_file_url = f'static/excel/{merged_filename}'
        
        total_rows = merged_sheet.max_row
        
        # 返回处理结果
        return jsonify({
            'success': True,
            'file_path': merged_file_url,
            'original_sheets': original_sheets,
            'total_rows': total_rows,
            'deleted_rows': delete_count
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'文件处理失败: {str(e)}'
        }), 500
    finally:
        # 清理临时文件
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)


if __name__ == '__main__':
    with app.app_context():
        # 先创建所有表（删除旧文件后）
        db.create_all()

        # 尝试更新表结构（兼容已有数据的情况）
        try:
            update_database_schema()
        except:
            pass

        # 添加管理员账号
        if not User.query.filter_by(username='111').first():
            # 管理员
            admin = User(username='111', password='111', is_admin=True)
            db.session.add(admin)
            db.session.commit()
    app.run(debug=True)